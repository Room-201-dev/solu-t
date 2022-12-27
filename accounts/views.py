import openpyxl
from django.shortcuts import render, redirect
from django.conf import settings
from django.core.mail import EmailMessage
from django.urls import reverse_lazy
from allauth.account import views
from django.views.generic import View
from django.views.generic.edit import FormView
from django.contrib.auth import login
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import LoginView, LogoutView
import datetime
from dateutil.relativedelta import relativedelta
from django.http import HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.mail import send_mail
from django.db.models import Q
from functools import reduce
from operator import and_

from solu_t.models import CustomUser, Notice, ApplyList, ApplyData, ContactData

from .forms import NoticePostForm, SuperUserLoginForm
from solu_t.forms import ContactForm

from django.views.decorators.csrf import requires_csrf_token
from django.http import HttpResponseServerError


# Create your views here.


class LoginManagerView(LoginView):
    def post(self, request, *args, **kwargs):
        form = SuperUserLoginForm(data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            user = CustomUser.objects.get(username=username)
            login(request, user)
            return redirect('manager_page')
        return render(request, 'accounts/login.html', {'form': form})

    def get(self, request, *args, **kwargs):
        form = SuperUserLoginForm(request.POST)
        return render(request, 'accounts/login.html', {'form': form})


class LogoutManagerView(views.LogoutView):
    template_name = 'accounts/logout.html'


class ManagerPageView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        incomplete_contact = ContactData.objects.order_by('-id')
        return render(request, 'accounts/admin_page.html', {
            'incomplete_contact': incomplete_contact
        })


class PostNoticeView(LoginRequiredMixin, FormView):
    template_name = 'accounts/post_notice_form.html'
    form_class = NoticePostForm
    success_url = reverse_lazy('notice_list')

    def post(self, request, *args, **kwargs):
        form = NoticePostForm(request.POST or None)

        if form.is_valid():
            notice_data = Notice()
            notice_data.title = form.cleaned_data['title']
            notice_data.content = form.cleaned_data['content']
            notice_data.important = form.cleaned_data['choice_important']
            notice_data.base = form.cleaned_data['choice_base']
            notice_data.shift = form.cleaned_data['choice_shift']
            notice_data.tag = form.cleaned_data['choice_tag']

            notice_data.save()

            subject = form.cleaned_data['title']
            content = form.cleaned_data['content']
            from_email = settings.DEFAULT_FROM_EMAIL
            bcc = []

            if form.cleaned_data['choice_shift'] == '全体':
                for mail_push in CustomUser.objects.filter(base=form.cleaned_data['choice_base']):
                    bcc.append(mail_push.email)
            elif form.cleaned_data['choice_shift'] == '日勤' or form.cleaned_data['choice_shift'] == '日勤':
                for mail_push in CustomUser.objects.filter(shift=form.cleaned_data['choice_shift'],
                                                           base=form.cleaned_data['choice_base']):
                    bcc.append(mail_push.email)

            email_content = '※このメールは送信専用アドレスから配信されています。\n\n' + content + '\n\nhttps://solu-t.herokuapp.com/'

            email = EmailMessage(subject, email_content, from_email, [], bcc)
            email.send()
            return redirect('notice_list')

        return render(request, 'accounts/post_notice_form.html', {
            'form': form,
        })


class NoticeListView(View):
    def get(self, request, *args, **kwargs):
        notice_list = Notice.objects.order_by('-id')
        return render(request, 'accounts/admin_notice_list.html', {
            'notice_list': notice_list
        })


class NoticeDetailView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        notice_detail = Notice.objects.get(id=self.kwargs['pk'])
        return render(request, 'accounts/admin_notice_detail.html', {
            'notice_detail': notice_detail
        })


class NoticeEditView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        notice_detail = Notice.objects.get(id=self.kwargs['pk'])
        notice_edit_form = NoticePostForm(
            request.POST or None,
            initial={
                'choice_base': notice_detail.base,
                'choice_shift': notice_detail.shift,
                'choice_important': notice_detail.important,
                'title': notice_detail.title,
                'content': notice_detail.content,
            }
        )
        return render(request, 'accounts/admin_notice_edit.html', {
            'notice_edit_form': notice_edit_form,
        })

    def post(self, request, *args, **kwargs):
        form = NoticePostForm(request.POST or None)

        if form.is_valid():
            notice_data = Notice.objects.get(id=self.kwargs['pk'])
            notice_data.title = form.cleaned_data['title']
            notice_data.content = form.cleaned_data['content']
            notice_data.important = form.cleaned_data['choice_important']
            notice_data.base = form.cleaned_data['choice_base']
            notice_data.shift = form.cleaned_data['choice_shift']
            notice_data.tag = form.cleaned_data['choice_tag']

            notice_data.save()

            subject = form.cleaned_data['title']
            content = form.cleaned_data['content']
            from_email = settings.DEFAULT_FROM_EMAIL
            bcc = []

            if form.cleaned_data['choice_shift'] == '全体':
                for mail_push in CustomUser.objects.filter(base=form.cleaned_data['choice_base']):
                    bcc.append(mail_push.email)
            elif form.cleaned_data['choice_shift'] == '日勤' or form.cleaned_data['choice_shift'] == '日勤':
                for mail_push in CustomUser.objects.filter(shift=form.cleaned_data['choice_shift'],
                                                           base=form.cleaned_data['choice_base']):
                    bcc.append(mail_push.email)

            email_content = '※このメールは送信専用アドレスから配信されています。\n\n' + content + '\n\nhttps://solu-t.herokuapp.com/'

            email = EmailMessage(subject, email_content, from_email, [], bcc)
            email.send()
            return redirect('notice_list')

        return render(request, 'accounts/admin_notice_edit.html', {
            'form': form,
        })


class NoticeDeleteView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        notice_data = Notice.objects.get(id=self.kwargs['pk'])

        return render(request, 'accounts/notice_delete.html', {
            'notice_data': notice_data
        })

    def post(self, request, *args, **kwargs):
        notice_data = Notice.objects.get(id=self.kwargs['pk'])
        notice_data.delete()
        return redirect('notice_list')


class ApplyListView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        apply_data = ApplyData.objects.all()

        return render(request, 'accounts/admin_apply_data.html', {
            'apply_data': apply_data
        })

    def post(self, request, *args, **kwargs):
        records = ApplyData.objects.all()
        records.delete()

        return redirect('apply_data')


def export(request):
    apply_list = ApplyData.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=staff_leave_list.xlsx'

    ws.cell(2, 2).value = '所属拠点'
    ws.cell(2, 3).value = '氏名'
    ws.cell(2, 4).value = '申請内容'
    ws.cell(2, 5).value = '申請日'
    i = 3

    for apply in apply_list:
        ws.cell(i, 2).value = apply.base
        ws.cell(i, 3).value = apply.name
        ws.cell(i, 4).value = apply.choice_kind
        if apply.date is None:
            ws.cell(i, 5).value = apply.refresh_date
        else:
            ws.cell(i, 5).value = apply.date
        i += 1

    wb.save(response)

    return response


class ApplyTYO4ThisMonthView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        today = datetime.date.today()
        first_day = today + relativedelta(day=1)
        last_day = today + relativedelta(months=+1, day=1, days=-1)
        date_this_month = ApplyList.objects.filter(date__range=(first_day, last_day)).filter(base="青梅")
        holiday = ApplyList.objects.filter(holiday_date__range=(first_day, last_day)).filter(base="青梅")
        workday = ApplyList.objects.filter(work_date__range=(first_day, last_day)).filter(base="青梅")
        plus_work = ApplyList.objects.filter(plus_work__range=(first_day, last_day)).filter(base="青梅")
        refresh_date = ApplyList.objects.filter(refresh_date__range=(first_day, last_day)).filter(base="青梅")
        early_date = ApplyList.objects.filter(early_date__range=(first_day, last_day)).filter(base='青梅')
        this_month = date_this_month.union(holiday, workday, plus_work, refresh_date, early_date)

        return render(request, 'accounts/admin_applytyo4_thismonth.html', {
            'this_month': this_month,
        })

    def post(self, request, *args, **kwargs):
        post_pks = request.POST.getlist('delete')
        for send_e in ApplyList.objects.filter(pk__in=post_pks):
            name = send_e.name.replace('　', '')
            email = send_e.email
            req = send_e.choice_kind.replace(' ', '')

            if send_e.date:
                req_date = send_e.date
            elif send_e.work_date:
                req_date = send_e.work_date
            elif send_e.plus_work:
                req_date = send_e.plus_work
            elif send_e.refresh_date:
                req_date = send_e.refresh_date
            elif send_e.early_date:
                req_date = send_e.early_date

            if send_e.remarks_area:
                paid_leave = send_e.remarks_area
                req = ''
            else:
                paid_leave = ''

            subject = '申請が完了いたしました'
            content = '※このメールは送信専用アドレスから配信されています。\n\n{name} 様\n\nお疲れ様です。\nシフト担当でございます。\n\n頂戴しておりました {req_date}の{req}{paid_leave} 処理が完了いたしました。\n\n引き続きよろしくお願いいたします。\n\nhttps://solu-t.herokuapp.com/'.format(
                name=name, req=req, req_date=req_date, paid_leave=paid_leave)
            recipient_list = [email]
            complete_mail = EmailMessage(subject, content, 'towa-cast@complate', recipient_list)
            complete_mail.send()
        ApplyList.objects.filter(pk__in=post_pks).delete()
        return redirect('apply_tyo4_thismonth')


class ApplyTYO4NextMonthView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        today = datetime.date.today()
        first_day = today + relativedelta(months=+1, day=1)
        last_day = today + relativedelta(months=+2, day=1, days=-1)
        date_next_month = ApplyList.objects.filter(date__range=(first_day, last_day)).filter(base="青梅")
        holiday = ApplyList.objects.filter(holiday_date__range=(first_day, last_day)).filter(base="青梅")
        workday = ApplyList.objects.filter(work_date__range=(first_day, last_day)).filter(base="青梅")
        plus_work = ApplyList.objects.filter(plus_work__range=(first_day, last_day)).filter(base="青梅")
        refresh_date = ApplyList.objects.filter(refresh_date__range=(first_day, last_day)).filter(base="青梅")
        early_date = ApplyList.objects.filter(early_date__range=(first_day, last_day)).filter(base='青梅')
        next_month = date_next_month.union(holiday, workday, plus_work, refresh_date, early_date)

        return render(request, 'accounts/admin_applytyo4_nextmonth.html', {
            'next_month': next_month,
        })

    def post(self, request, *args, **kwargs):
        post_pks = request.POST.getlist('delete')
        for send_e in ApplyList.objects.filter(pk__in=post_pks):
            name = send_e.name.replace('　', '')
            email = send_e.email
            req = send_e.choice_kind.replace(' ', '')

            if send_e.date:
                req_date = send_e.date
            elif send_e.work_date:
                req_date = send_e.work_date
            elif send_e.plus_work:
                req_date = send_e.plus_work
            elif send_e.refresh_date:
                req_date = send_e.refresh_date
            elif send_e.early_date:
                req_date = send_e.early_date

            if send_e.remarks_area:
                paid_leave = send_e.remarks_area
                req = ''
            else:
                paid_leave = ''

            subject = '申請が完了いたしました'
            content = '※このメールは送信専用アドレスから配信されています。\n\n{name} 様\n\nお疲れ様です。\nシフト担当でございます。\n\n頂戴しておりました {req_date}の{req}{paid_leave} 処理が完了いたしました。\n\n引き続きよろしくお願いいたします。\n\nhttps://solu-t.herokuapp.com/'.format(
                name=name, req=req, req_date=req_date, paid_leave=paid_leave)
            recipient_list = [email]
            complete_mail = EmailMessage(subject, content, 'towa-cast@complate', recipient_list)
            complete_mail.send()
        ApplyList.objects.filter(pk__in=post_pks).delete()
        return redirect('apply_tyo4_nextmonth')


class ApplyTYO4TwoMonthView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        today = datetime.date.today()
        first_day = today + relativedelta(months=+2, day=1)
        last_day = today + relativedelta(months=+3, day=1, days=-1)
        date_next_month = ApplyList.objects.filter(date__range=(first_day, last_day)).filter(base="青梅")
        holiday = ApplyList.objects.filter(holiday_date__range=(first_day, last_day)).filter(base="青梅")
        workday = ApplyList.objects.filter(work_date__range=(first_day, last_day)).filter(base="青梅")
        plus_work = ApplyList.objects.filter(plus_work__range=(first_day, last_day)).filter(base="青梅")
        refresh_date = ApplyList.objects.filter(refresh_date__range=(first_day, last_day)).filter(base="青梅")
        early_date = ApplyList.objects.filter(early_date__range=(first_day, last_day)).filter(base='青梅')
        two_month = date_next_month.union(holiday, workday, plus_work, refresh_date, early_date)

        return render(request, 'accounts/admin_applytyo4_twomonth.html', {
            'two_month': two_month,
        })

    def post(self, request, *args, **kwargs):
        post_pks = request.POST.getlist('delete')
        for send_e in ApplyList.objects.filter(pk__in=post_pks):
            name = send_e.name.replace('　', '')
            email = send_e.email
            req = send_e.choice_kind.replace(' ', '')

            if send_e.date:
                req_date = send_e.date
            elif send_e.work_date:
                req_date = send_e.work_date
            elif send_e.plus_work:
                req_date = send_e.plus_work
            elif send_e.refresh_date:
                req_date = send_e.refresh_date
            elif send_e.early_date:
                req_date = send_e.early_date

            if send_e.remarks_area:
                paid_leave = send_e.remarks_area
                req = ''
            else:
                paid_leave = ''

            subject = '申請が完了いたしました'
            content = '※このメールは送信専用アドレスから配信されています。\n\n{name} 様\n\nお疲れ様です。\nシフト担当でございます。\n\n頂戴しておりました {req_date}の{req}{paid_leave} 処理が完了いたしました。\n\n引き続きよろしくお願いいたします。\n\nhttps://solu-t.herokuapp.com/'.format(
                name=name, req=req, req_date=req_date, paid_leave=paid_leave)
            recipient_list = [email]
            complete_mail = EmailMessage(subject, content, 'towa-cast@complate', recipient_list)
            complete_mail.send()
        ApplyList.objects.filter(pk__in=post_pks).delete()
        return redirect('apply_tyo4_twomonth')


class ApplyTYO6ThisMonthView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        keyword = request.GET.get('keyword')

        today = datetime.date.today()
        first_day = today + relativedelta(day=1)
        last_day = today + relativedelta(months=+1, day=1, days=-1)
        date_this_month = ApplyList.objects.filter(date__range=(first_day, last_day)).filter(base="坂戸")
        holiday = ApplyList.objects.filter(holiday_date__range=(first_day, last_day)).filter(base="坂戸")
        workday = ApplyList.objects.filter(work_date__range=(first_day, last_day)).filter(base="坂戸")
        plus_work = ApplyList.objects.filter(plus_work__range=(first_day, last_day)).filter(base="坂戸")
        refresh_date = ApplyList.objects.filter(refresh_date__range=(first_day, last_day)).filter(base="坂戸")
        early_date = ApplyList.objects.filter(early_date__range=(first_day, last_day)).filter(base='坂戸')
        this_month = date_this_month.union(holiday, workday, plus_work, refresh_date, early_date)

        return render(request, 'accounts/admin_applytyo6_thismonth.html', {
            'this_month': this_month,
        })

    def post(self, request, *args, **kwargs):
        post_pks = request.POST.getlist('delete')
        for send_e in ApplyList.objects.filter(pk__in=post_pks):
            name = send_e.name.replace('　', '')
            email = send_e.email
            req = send_e.choice_kind.replace(' ', '')

            if send_e.date:
                req_date = send_e.date
            elif send_e.work_date:
                req_date = send_e.work_date
            elif send_e.plus_work:
                req_date = send_e.plus_work
            elif send_e.refresh_date:
                req_date = send_e.refresh_date
            elif send_e.early_date:
                req_date = send_e.early_date

            if send_e.remarks_area:
                paid_leave = send_e.remarks_area
                req = ''
            else:
                paid_leave = ''

            subject = '申請が完了いたしました'
            content = '※このメールは送信専用アドレスから配信されています。\n\n{name} 様\n\nお疲れ様です。\nシフト担当でございます。\n\n頂戴しておりました {req_date}の{req}{paid_leave} 処理が完了いたしました。\n\n引き続きよろしくお願いいたします。\n\nhttps://solu-t.herokuapp.com/'.format(
                name=name, req=req, req_date=req_date, paid_leave=paid_leave)
            recipient_list = [email]
            complete_mail = EmailMessage(subject, content, 'towa-cast@complate', recipient_list)
            complete_mail.send()
        ApplyList.objects.filter(pk__in=post_pks).delete()
        return redirect('apply_tyo6_thismonth')


class ApplyTYO6NextMonthView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        today = datetime.date.today()
        first_day = today + relativedelta(months=+1, day=1)
        last_day = today + relativedelta(months=+2, day=1, days=-1)
        date_next_month = ApplyList.objects.filter(date__range=(first_day, last_day)).filter(base="坂戸")
        holiday = ApplyList.objects.filter(holiday_date__range=(first_day, last_day)).filter(base="坂戸")
        workday = ApplyList.objects.filter(work_date__range=(first_day, last_day)).filter(base="坂戸")
        plus_work = ApplyList.objects.filter(plus_work__range=(first_day, last_day)).filter(base="坂戸")
        refresh_date = ApplyList.objects.filter(refresh_date__range=(first_day, last_day)).filter(base="坂戸")
        early_date = ApplyList.objects.filter(early_date__range=(first_day, last_day)).filter(base='坂戸')
        next_month = date_next_month.union(holiday, workday, plus_work, refresh_date, early_date)

        return render(request, 'accounts/admin_applytyo6_nextmonth.html', {
            'next_month': next_month,
        })

    def post(self, request, *args, **kwargs):
        post_pks = request.POST.getlist('delete')
        for send_e in ApplyList.objects.filter(pk__in=post_pks):
            name = send_e.name.replace('　', '')
            email = send_e.email
            req = send_e.choice_kind.replace(' ', '')

            if send_e.date:
                req_date = send_e.date
            elif send_e.work_date:
                req_date = send_e.work_date
            elif send_e.plus_work:
                req_date = send_e.plus_work
            elif send_e.refresh_date:
                req_date = send_e.refresh_date
            elif send_e.early_date:
                req_date = send_e.early_date

            if send_e.remarks_area:
                paid_leave = send_e.remarks_area
                req = ''
            else:
                paid_leave = ''

            subject = '申請が完了いたしました'
            content = '※このメールは送信専用アドレスから配信されています。\n\n{name} 様\n\nお疲れ様です。\nシフト担当でございます。\n\n頂戴しておりました {req_date}の{req}{paid_leave} 処理が完了いたしました。\n\n引き続きよろしくお願いいたします。\n\nhttps://solu-t.herokuapp.com/'.format(
                name=name, req=req, req_date=req_date, paid_leave=paid_leave)
            recipient_list = [email]
            complete_mail = EmailMessage(subject, content, 'towa-cast@complate', recipient_list)
            complete_mail.send()
        ApplyList.objects.filter(pk__in=post_pks).delete()
        return redirect('apply_tyo6_nextmonth')


class ApplyTYO6TwoMonthView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        today = datetime.date.today()
        first_day = today + relativedelta(months=+2, day=1)
        last_day = today + relativedelta(months=+3, day=1, days=-1)
        date_next_month = ApplyList.objects.filter(date__range=(first_day, last_day)).filter(base="坂戸")
        holiday = ApplyList.objects.filter(holiday_date__range=(first_day, last_day)).filter(base="坂戸")
        workday = ApplyList.objects.filter(work_date__range=(first_day, last_day)).filter(base="坂戸")
        plus_work = ApplyList.objects.filter(plus_work__range=(first_day, last_day)).filter(base="坂戸")
        refresh_date = ApplyList.objects.filter(refresh_date__range=(first_day, last_day)).filter(base="坂戸")
        early_date = ApplyList.objects.filter(early_date__range=(first_day, last_day)).filter(base='坂戸')
        two_month = date_next_month.union(holiday, workday, plus_work, refresh_date, early_date)

        return render(request, 'accounts/admin_applytyo6_twomonth.html', {
            'two_month': two_month,
        })

    def post(self, request, *args, **kwargs):
        post_pks = request.POST.getlist('delete')
        for send_e in ApplyList.objects.filter(pk__in=post_pks):
            name = send_e.name.replace('　', '')
            email = send_e.email
            req = send_e.choice_kind.replace(' ', '')

            if send_e.date:
                req_date = send_e.date
            elif send_e.work_date:
                req_date = send_e.work_date
            elif send_e.plus_work:
                req_date = send_e.plus_work
            elif send_e.refresh_date:
                req_date = send_e.refresh_date
            elif send_e.early_date:
                req_date = send_e.early_date

            if send_e.remarks_area:
                paid_leave = send_e.remarks_area
                req = ''
            else:
                paid_leave = ''

            subject = '申請が完了いたしました'
            content = '※このメールは送信専用アドレスから配信されています。\n\n{name} 様\n\nお疲れ様です。\nシフト担当でございます。\n\n頂戴しておりました {req_date}の{req}{paid_leave} 処理が完了いたしました。\n\n引き続きよろしくお願いいたします。\n\nhttps://solu-t.herokuapp.com/'.format(
                name=name, req=req, req_date=req_date, paid_leave=paid_leave)
            recipient_list = [email]
            complete_mail = EmailMessage(subject, content, 'towa-cast@complate', recipient_list)
            complete_mail.send()
        ApplyList.objects.filter(pk__in=post_pks).delete()
        return redirect('apply_tyo8_twomonth')


class ApplyTYO8ThisMonthView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        today = datetime.date.today()
        first_day = today + relativedelta(day=1)
        last_day = today + relativedelta(months=+1, day=1, days=-1)
        date_this_month = ApplyList.objects.filter(date__range=(first_day, last_day)).filter(base="相模原")
        holiday = ApplyList.objects.filter(holiday_date__range=(first_day, last_day)).filter(base="相模原")
        workday = ApplyList.objects.filter(work_date__range=(first_day, last_day)).filter(base="相模原")
        plus_work = ApplyList.objects.filter(plus_work__range=(first_day, last_day)).filter(base="相模原")
        refresh_date = ApplyList.objects.filter(refresh_date__range=(first_day, last_day)).filter(base="相模原")
        early_date = ApplyList.objects.filter(early_date__range=(first_day, last_day)).filter(base='相模原')
        this_month = date_this_month.union(holiday, workday, plus_work, refresh_date, early_date)

        return render(request, 'accounts/admin_applytyo8_thismonth.html', {
            'this_month': this_month,
        })

    def post(self, request, *args, **kwargs):
        post_pks = request.POST.getlist('delete')
        for send_e in ApplyList.objects.filter(pk__in=post_pks):
            name = send_e.name.replace('　', '')
            email = send_e.email
            req = send_e.choice_kind.replace(' ', '')

            if send_e.date:
                req_date = send_e.date
            elif send_e.work_date:
                req_date = send_e.work_date
            elif send_e.plus_work:
                req_date = send_e.plus_work
            elif send_e.refresh_date:
                req_date = send_e.refresh_date
            elif send_e.early_date:
                req_date = send_e.early_date

            if send_e.remarks_area:
                paid_leave = send_e.remarks_area
                req = ''
            else:
                paid_leave = ''

            subject = '申請が完了いたしました'
            content = '※このメールは送信専用アドレスから配信されています。\n\n{name} 様\n\nお疲れ様です。\nシフト担当でございます。\n\n頂戴しておりました {req_date}の{req}{paid_leave} 処理が完了いたしました。\n\n引き続きよろしくお願いいたします。\n\nhttps://solu-t.herokuapp.com/'.format(
                name=name, req=req, req_date=req_date, paid_leave=paid_leave)
            recipient_list = [email]
            complete_mail = EmailMessage(subject, content, 'towa-cast@complate', recipient_list)
            complete_mail.send()
        ApplyList.objects.filter(pk__in=post_pks).delete()
        return redirect('apply_tyo8_thismonth')


class ApplyTYO8NextMonthView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        today = datetime.date.today()
        first_day = today + relativedelta(months=+1, day=1)
        last_day = today + relativedelta(months=+2, day=1, days=-1)
        date_next_month = ApplyList.objects.filter(date__range=(first_day, last_day)).filter(base="相模原")
        holiday = ApplyList.objects.filter(holiday_date__range=(first_day, last_day)).filter(base="相模原")
        workday = ApplyList.objects.filter(work_date__range=(first_day, last_day)).filter(base="相模原")
        plus_work = ApplyList.objects.filter(plus_work__range=(first_day, last_day)).filter(base="相模原")
        refresh_date = ApplyList.objects.filter(refresh_date__range=(first_day, last_day)).filter(base="相模原")
        early_date = ApplyList.objects.filter(early_date__range=(first_day, last_day)).filter(base='相模原')
        next_month = date_next_month.union(holiday, workday, plus_work, refresh_date, early_date)

        return render(request, 'accounts/admin_applytyo8_nextmonth.html', {
            'next_month': next_month,
        })

    def post(self, request, *args, **kwargs):
        post_pks = request.POST.getlist('delete')
        for send_e in ApplyList.objects.filter(pk__in=post_pks):
            name = send_e.name.replace('　', '')
            email = send_e.email
            req = send_e.choice_kind.replace(' ', '')

            if send_e.date:
                req_date = send_e.date
            elif send_e.work_date:
                req_date = send_e.work_date
            elif send_e.plus_work:
                req_date = send_e.plus_work
            elif send_e.refresh_date:
                req_date = send_e.refresh_date
            elif send_e.early_date:
                req_date = send_e.early_date

            if send_e.remarks_area:
                paid_leave = send_e.remarks_area
                req = ''
            else:
                paid_leave = ''

            subject = '申請が完了いたしました'
            content = '※このメールは送信専用アドレスから配信されています。\n\n{name} 様\n\nお疲れ様です。\nシフト担当でございます。\n\n頂戴しておりました {req_date}の{req}{paid_leave} 処理が完了いたしました。\n\n引き続きよろしくお願いいたします。\n\nhttps://solu-t.herokuapp.com/'.format(
                name=name, req=req, req_date=req_date, paid_leave=paid_leave)
            recipient_list = [email]
            complete_mail = EmailMessage(subject, content, 'towa-cast@complate', recipient_list)
            complete_mail.send()
        ApplyList.objects.filter(pk__in=post_pks).delete()
        return redirect('apply_tyo8_nextmonth')


class ApplyTYO8TwoMonthView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        today = datetime.date.today()
        first_day = today + relativedelta(months=+2, day=1)
        last_day = today + relativedelta(months=+3, day=1, days=-1)
        date_next_month = ApplyList.objects.filter(date__range=(first_day, last_day)).filter(base="相模原")
        holiday = ApplyList.objects.filter(holiday_date__range=(first_day, last_day)).filter(base="相模原")
        workday = ApplyList.objects.filter(work_date__range=(first_day, last_day)).filter(base="相模原")
        plus_work = ApplyList.objects.filter(plus_work__range=(first_day, last_day)).filter(base="相模原")
        refresh_date = ApplyList.objects.filter(refresh_date__range=(first_day, last_day)).filter(base="相模原")
        early_date = ApplyList.objects.filter(early_date__range=(first_day, last_day)).filter(base='相模原')
        two_month = date_next_month.union(holiday, workday, plus_work, refresh_date, early_date)

        return render(request, 'accounts/admin_applytyo8_twomonth.html', {
            'two_month': two_month,
        })

    def post(self, request, *args, **kwargs):
        post_pks = request.POST.getlist('delete')
        for send_e in ApplyList.objects.filter(pk__in=post_pks):
            name = send_e.name.replace('　', '')
            email = send_e.email
            req = send_e.choice_kind.replace(' ', '')

            if send_e.date:
                req_date = send_e.date
            elif send_e.work_date:
                req_date = send_e.work_date
            elif send_e.plus_work:
                req_date = send_e.plus_work
            elif send_e.refresh_date:
                req_date = send_e.refresh_date
            elif send_e.early_date:
                req_date = send_e.early_date

            if send_e.remarks_area:
                paid_leave = send_e.remarks_area
                req = ''
            else:
                paid_leave = ''

            subject = '申請が完了いたしました'
            content = '※このメールは送信専用アドレスから配信されています。\n\n{name} 様\n\nお疲れ様です。\nシフト担当でございます。\n\n頂戴しておりました {req_date}の{req}{paid_leave} 処理が完了いたしました。\n\n引き続きよろしくお願いいたします。\n\nhttps://solu-t.herokuapp.com/'.format(
                name=name, req=req, req_date=req_date, paid_leave=paid_leave)
            recipient_list = [email]
            complete_mail = EmailMessage(subject, content, 'towa-cast@complate', recipient_list)
            complete_mail.send()
        ApplyList.objects.filter(pk__in=post_pks).delete()
        return redirect('apply_tyo8_twomonth')


class ContactListView(View):
    def get(self, request, *args, **kwargs):
        contact_list = ContactData.objects.order_by('-id')
        paginator = Paginator(contact_list, 20)
        contact_page = request.GET.get('page', 1)
        try:
            contact_page = paginator.page(contact_page)
        except PageNotAnInteger:
            contact_page = paginator.page(1)
        except EmptyPage:
            contact_page = paginator.page(1)
        context = {'contact_page': contact_page}
        return render(request, 'accounts/contact_list.html', context)


class ContactDetailView(View):
    def get(self, request, *args, **kwargs):
        contact_data = ContactData.objects.get(id=self.kwargs['pk'])

        return render(request, 'accounts/contact_detail.html', {
            'contact_data': contact_data
        })


class ContactReplyView(View):
    def get(self, request, *args, **kwargs):
        contact_data = ContactData.objects.get(id=self.kwargs['pk'])
        reply_form = ContactForm(
            request.POST or None,
            initial={
                'kind_contact': contact_data.contact_kind,
                'base': contact_data.base,
                'name': contact_data.name,
                'email': contact_data.email,
            }
        )

        return render(request, 'accounts/contact_reply.html', {
            'contact_data': contact_data,
            'reply_form': reply_form
        })

    def post(self, request, *args, **kwargs):
        reply_form = ContactForm(request.POST or None)
        user = request.user

        if reply_form.is_valid():
            contact_data = ContactData.objects.get(id=self.kwargs['pk'])
            contact_data.base = reply_form.cleaned_data['base']
            contact_data.contact_kind = reply_form.cleaned_data['kind_contact']
            contact_data.email = reply_form.cleaned_data['email']
            contact_data.name = reply_form.cleaned_data['name']
            contact_data.message = contact_data.message + '\n\n\n>>> 担当者：{user}\n{message}'.format(user=user.last_name,
                                                                                                   message=
                                                                                                   reply_form.cleaned_data[
                                                                                                       'message'])
            contact_data.tag = user.last_name + '返信済み'
            contact_data.save()

            subject = 'お問い合わせに返信がきました！'
            content = '※このメールは送信専用アドレスから配信されています。\n\n' + contact_data.message + '\n\nhttps://solu-t.herokuapp.com/'
            recipient_list = [reply_form.cleaned_data['email']]
            send_mail(subject, content, 'manager@towa', recipient_list, )

            return redirect('contact_list')

        else:
            return HttpResponse('無効なヘッダが検出されました')


class DeleteContactView(View):
    def get(self, request, *args, **kwargs):
        contact_data = ContactData.objects.get(id=self.kwargs['pk'])

        return render(request, 'accounts/delete_contact.html', {
            'contact_data': contact_data
        })

    def post(self, request, *args, **kwargs):
        contact_data = ContactData.objects.get(id=self.kwargs['pk'])
        contact_data.delete()
        return redirect('contact_list')


class SearchContactView(View):
    def get(self, request, *args, **kwargs):
        contact_page = ContactData.objects.order_by('-id')
        keyword = request.GET.get('keyword')

        if keyword:
            exclusion_list = set([' ', '　'])
            query_list = ''
            for word in keyword:
                if not word in exclusion_list:
                    query_list += word

            query = reduce(and_, [Q(base__icontains=q) | Q(name__icontains=q) for q in query_list])
            contact_page = contact_page.filter(query)

        return render(request, 'accounts/contact_list.html', {
            'keyword': keyword,
            'contact_page': contact_page
        })
